import streamlit as st
import pdfplumber
import pandas as pd
import re
import tempfile
import base64
import hashlib
import difflib
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# =============================
# 🎯 CONFIG
# =============================
st.set_page_config(page_title="Smart PDF Comparator", layout="wide")

# =============================
# 🔐 LOGIN SYSTEM
# =============================
def check_login():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    if not st.session_state.logged_in:
        st.markdown("""
        <div style="display:flex; justify-content:center; margin-top:80px;">
        <div style="background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
                    padding: 2.5rem; border-radius: 20px; width:380px; text-align:center;">
            <h1 style="color:#e94560; margin-bottom:0.2rem;">🏭 Smart PDF</h1>
            <h2 style="color:#e94560; margin-top:0;">Comparator</h2>
            <p style="color:#a8b2d8; margin-bottom:2rem;">Please login to continue</p>
        </div>
        </div>
        """, unsafe_allow_html=True)

        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            username = st.text_input("👤 Username")
            password = st.text_input("🔑 Password", type="password")

            if st.button("🔓 Login", use_container_width=True):
                if username == "dreamy" and password == "YDreamy":
                    st.session_state.logged_in = True
                    st.rerun()
                else:
                    st.error("❌ Incorrect username or password")
        st.stop()

check_login()

# =============================
# 🧠 DEBUG MODE
# =============================
st.sidebar.title("⚙️ Settings")
debug_mode = st.sidebar.checkbox("🐞 Debug mode", value=False)

if st.sidebar.button("🚪 Logout"):
    st.session_state.logged_in = False
    st.rerun()

# =============================
# 📄 VIEW PDF
# =============================
def show_pdf(file):
    file.seek(0)
    data = file.read()
    st.download_button(
        label="📥 Télécharger / Ouvrir le PDF",
        data=data,
        file_name=file.name,
        mime="application/pdf"
    )

# =============================
# 📑 INDEX
# =============================
@st.cache_data
def extract_index_and_info(pdf_path):
    index_map = {}
    project_name = "Projet inconnu"

    with pdfplumber.open(pdf_path) as pdf:
        if len(pdf.pages) < 2:
            return {}, "", project_name

        text = pdf.pages[1].extract_text() or ""

        for line in text.split("\n"):
            if "project name" in line.lower():
                project_name = line.replace("Project Name", "").strip()

            if re.match(r"^\s*\d+\s+.+", line):
                match = re.match(r"^(\d+)\s+(.+)", line)
                if match:
                    index_map[match.group(2)] = int(match.group(1))

    return index_map, text, project_name

# =============================
# 🔥 FILTER INDEX
# =============================
@st.cache_data
def filter_sections(index_map):
    return {
        k: v for k, v in index_map.items()
        if "cover" not in k.lower() and "index" not in k.lower()
    }

# =============================
# 🔥 PAGE RANGE
# =============================
@st.cache_data
def get_section_ranges(index_map, total_pages):
    sorted_items = sorted(index_map.items(), key=lambda x: x[1])
    ranges = {}

    for i, (section, start) in enumerate(sorted_items):
        if i < len(sorted_items) - 1:
            end = sorted_items[i + 1][1] - 1
        else:
            end = total_pages

        ranges[section] = (start, end)

    return ranges

# =============================
# ⚡ PAGE DATA
# =============================
@st.cache_data
def get_pages_data(pdf_path):
    pages = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            clean = re.sub(r"\s+", " ", text).strip().lower()
            h = hashlib.md5(clean.encode()).hexdigest()
            pages.append((clean, h))

    return pages

# =============================
# ⚡ QUICK COMPARE
# =============================
def quick_compare(index_new, index_old, path_new, path_old):

    new_filtered = filter_sections(index_new)
    old_filtered = filter_sections(index_old)

    if set(new_filtered.keys()) != set(old_filtered.keys()):
        return False, {"reason": "Different structure (sections)"}

    for k in new_filtered:
        if new_filtered[k] != old_filtered[k]:
            return False, {"reason": f"Different page mapping in section: {k}"}

    new_pages = get_pages_data(path_new)
    old_pages = get_pages_data(path_old)

    min_len = min(len(new_pages), len(old_pages))
    diff_pages = []
    scores = []

    for i in range(min_len):
        txt1 = new_pages[i][0]
        txt2 = old_pages[i][0]
        ratio = difflib.SequenceMatcher(None, txt1, txt2).ratio()
        scores.append(ratio)
        if ratio < 0.99:
            diff_pages.append(i + 1)

    extra_pages = abs(len(new_pages) - len(old_pages))
    if extra_pages > 0:
        diff_pages.extend(list(range(min_len + 1, min_len + extra_pages + 1)))

    avg_similarity = sum(scores) / len(scores) if scores else 0
    similarity = round(avg_similarity * 100, 2)
    difference = round(100 - similarity, 2)

    if len(diff_pages) == 0:
        return True, {"reason": "Fully identical", "similarity": 100, "difference": 0}

    return False, {
        "reason": "Same structure but DIFFERENT CONTENT",
        "similarity": similarity,
        "difference": difference,
        "different_pages": diff_pages[:10]
    }

# =============================
# 🧠 CLEAN META ROWS
# =============================
def is_meta_row(row):
    text = " ".join([str(x) for x in row if pd.notna(x)]).lower()
    keywords = ["project", "client", "rev", "date", "page", "document", "sheet", "title"]
    return any(k in text for k in keywords)

# =============================
# 🧠 CLEAN COLUMNS
# =============================
def clean_columns(cols):
    seen = {}
    clean = []

    for c in cols:
        c = str(c).strip()
        if c == "" or c.lower() == "nan":
            c = "col"
        if c in seen:
            seen[c] += 1
            clean.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            clean.append(c)

    return clean

# =============================
# 🔑 HEADER DETECTION
# =============================
def detect_real_header(pdf_path, section_start_page, nb_cols):
    with pdfplumber.open(pdf_path) as pdf:
        page_idx = section_start_page - 1
        if page_idx >= len(pdf.pages):
            return None

        tables = pdf.pages[page_idx].extract_tables()
        if not tables:
            return None

        rows = [[str(c).strip() if c else "" for c in row] for row in tables[0]]

        if len(rows) < 3:
            return None

        data_start = None
        for i, r in enumerate(rows):
            if re.match(r"^\d+$", str(r[0]).strip()):
                data_start = i
                break

        if data_start is None or data_start < 2:
            return None

        h1 = rows[data_start - 2]
        h2 = rows[data_start - 1]

        max_len = max(len(h1), len(h2), nb_cols)
        h1 += [""] * (max_len - len(h1))
        h2 += [""] * (max_len - len(h2))

        merged = []
        for a, b in zip(h1[:nb_cols], h2[:nb_cols]):
            a, b = a.strip(), b.strip()
            merged.append(f"{a} {b}".strip() if a and b else a or b or "col")

        return merged

# =============================
# 📊 EXTRACT TABLES
# =============================
@st.cache_data
def extract_tables_range(pdf_path, start, end):
    tables = []

    with pdfplumber.open(pdf_path) as pdf:
        for i in range(start - 1, end):
            for t in pdf.pages[i].extract_tables():
                df = pd.DataFrame(t)
                df.dropna(how="all", inplace=True)
                df = df.map(lambda x: str(x).strip() if pd.notna(x) else x)
                tables.append(df)

    return tables

# =============================
# 🧹 MERGE CLEAN
# =============================
def merge_tables_clean(tables, pdf_path, section_start_page):

    if not tables:
        return pd.DataFrame()

    nb_cols = next((len(df.columns) for df in tables if not df.empty), None)
    if nb_cols is None:
        return pd.DataFrame()

    header = detect_real_header(pdf_path, section_start_page, nb_cols)
    header = clean_columns(header) if header else [f"Col{i+1}" for i in range(nb_cols)]

    final = []

    for df in tables:
        df = df.dropna(how="all").reset_index(drop=True)
        if df.empty:
            continue

        df = df[~df.apply(is_meta_row, axis=1)]
        df = df[df.iloc[:, 0].astype(str).str.match(r"^\d+$", na=False)]

        if df.empty:
            continue

        if len(df.columns) < len(header):
            for i in range(len(header) - len(df.columns)):
                df[f"_extra_{i}"] = ""
        else:
            df = df.iloc[:, :len(header)]

        df.columns = header
        final.append(df)

    return pd.concat(final, ignore_index=True) if final else pd.DataFrame()

# =============================
# 🔬 DEEP COMPARISON ENGINE
# =============================
def deep_compare(df_new, df_old, key_col, selected_cols):

    df_new = df_new.fillna("")
    df_old = df_old.fillna("")

    df_new = df_new.set_index(key_col)
    df_old = df_old.set_index(key_col)

    new_keys = set(df_new.index)
    old_keys = set(df_old.index)

    added   = list(new_keys - old_keys)
    deleted = list(old_keys - new_keys)
    common  = list(new_keys & old_keys)

    modified = {}

    for k in common:
        changes = {}
        for col in selected_cols:
            if col in df_new.columns and col in df_old.columns:
                v1 = str(df_old.loc[k, col])
                v2 = str(df_new.loc[k, col])
                if v1 != v2:
                    changes[col] = {"old": v1, "new": v2}
        if changes:
            modified[k] = changes

    return {"added": added, "deleted": deleted, "modified": modified}


# =============================
# 📝 GENERATE PDF REPORT
# =============================
def generate_pdf_report(result, section, key_col):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    # ── Colors ───────────────────────────────────────────────
    dark_blue   = colors.HexColor("#1a1a2e")
    accent_red  = colors.HexColor("#e94560")
    green       = colors.HexColor("#1e8449")
    red         = colors.HexColor("#c0392b")
    yellow_bg   = colors.HexColor("#fef9e7")
    green_bg    = colors.HexColor("#eafaf1")
    red_bg      = colors.HexColor("#fdf2f2")
    light_gray  = colors.HexColor("#f8f9fa")
    mid_gray    = colors.HexColor("#dee2e6")

    # ── Styles ───────────────────────────────────────────────
    styles = getSampleStyleSheet()

    style_title = ParagraphStyle("title",
        fontSize=28, textColor=accent_red, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=6)

    style_subtitle = ParagraphStyle("subtitle",
        fontSize=13, textColor=colors.HexColor("#a8b2d8"),
        alignment=TA_CENTER, fontName="Helvetica", spaceAfter=4)

    style_meta = ParagraphStyle("meta",
        fontSize=10, textColor=colors.HexColor("#6c757d"),
        alignment=TA_CENTER, fontName="Helvetica", spaceAfter=3)

    style_section_title = ParagraphStyle("sec_title",
        fontSize=14, textColor=dark_blue, fontName="Helvetica-Bold",
        spaceBefore=16, spaceAfter=6, borderPad=4)

    style_body = ParagraphStyle("body",
        fontSize=9, textColor=colors.HexColor("#333333"),
        fontName="Helvetica", spaceAfter=4, leading=14)

    style_footer = ParagraphStyle("footer",
        fontSize=8, textColor=colors.HexColor("#adb5bd"),
        alignment=TA_CENTER, fontName="Helvetica")

    now = datetime.datetime.now()
    now_str = now.strftime("%d/%m/%Y %H:%M")

    added    = result["added"]
    deleted  = result["deleted"]
    modified = result["modified"]
    n_total  = len(added) + len(deleted) + len(modified)

    story = []

    # ═══════════════════════════════════════════════════════
    # COVER PAGE
    # ═══════════════════════════════════════════════════════
    story.append(Spacer(1, 3*cm))

    # Title block
    cover_data = [[
        Paragraph("🏭 Smart PDF Comparator", style_title),
    ]]
    cover_table = Table(cover_data, colWidths=[16*cm])
    cover_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), dark_blue),
        ("ROUNDEDCORNERS", [10]),
        ("TOPPADDING",    (0,0), (-1,-1), 24),
        ("BOTTOMPADDING", (0,0), (-1,-1), 24),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 12),
    ]))
    story.append(cover_table)
    story.append(Spacer(1, 0.6*cm))

    story.append(Paragraph("Comparison Report / Rapport de Comparaison", style_subtitle))
    story.append(Spacer(1, 1.5*cm))

    # Info table
    info_data = [
        ["📋 Section",          section],
        ["🔑 Key Column",       key_col],
        ["📅 Date / Date",      now_str],
        ["⚙️ Generated by",    "Smart PDF Comparator"],
    ]
    info_table = Table(info_data, colWidths=[5*cm, 11*cm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (0,-1), dark_blue),
        ("TEXTCOLOR",    (0,0), (0,-1), colors.white),
        ("BACKGROUND",   (1,0), (1,-1), light_gray),
        ("TEXTCOLOR",    (1,0), (1,-1), colors.HexColor("#333333")),
        ("FONTNAME",     (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTNAME",     (1,0), (1,-1), "Helvetica"),
        ("FONTSIZE",     (0,0), (-1,-1), 10),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
        ("LEFTPADDING",  (0,0), (-1,-1), 10),
        ("GRID",         (0,0), (-1,-1), 0.5, mid_gray),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 2*cm))

    # KPI cards row
    kpi_data = [[
        Paragraph(f'<font color="#1e8449"><b>➕ Added</b></font><br/><font size="22" color="#1e8449"><b>{len(added)}</b></font><br/><font size="8" color="#6c757d">Ajoutés</font>', style_body),
        Paragraph(f'<font color="#c0392b"><b>➖ Deleted</b></font><br/><font size="22" color="#c0392b"><b>{len(deleted)}</b></font><br/><font size="8" color="#6c757d">Supprimés</font>', style_body),
        Paragraph(f'<font color="#d68910"><b>✏️ Modified</b></font><br/><font size="22" color="#d68910"><b>{len(modified)}</b></font><br/><font size="8" color="#6c757d">Modifiés</font>', style_body),
        Paragraph(f'<font color="#1a5276"><b>🔢 Total</b></font><br/><font size="22" color="#1a5276"><b>{n_total}</b></font><br/><font size="8" color="#6c757d">Total</font>', style_body),
    ]]
    kpi_table = Table(kpi_data, colWidths=[4*cm]*4)
    kpi_table.setStyle(TableStyle([
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0), (-1,-1), 14),
        ("BOTTOMPADDING",(0,0), (-1,-1), 14),
        ("BACKGROUND",   (0,0), (0,0), green_bg),
        ("BACKGROUND",   (1,0), (1,0), red_bg),
        ("BACKGROUND",   (2,0), (2,0), yellow_bg),
        ("BACKGROUND",   (3,0), (3,0), colors.HexColor("#eaf2fb")),
        ("GRID",         (0,0), (-1,-1), 0.5, mid_gray),
        ("ROUNDEDCORNERS", [6]),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 1.5*cm))

    story.append(HRFlowable(width="100%", thickness=1, color=accent_red))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Smart PDF Comparator  |  {now_str}", style_footer))

    # ═══════════════════════════════════════════════════════
    # PAGE 2 — ADDED EQUIPMENT
    # ═══════════════════════════════════════════════════════
    from reportlab.platypus import PageBreak
    story.append(PageBreak())

    story.append(Paragraph("➕ Added Equipment / Équipements Ajoutés", style_section_title))
    story.append(HRFlowable(width="100%", thickness=1.5, color=green))
    story.append(Spacer(1, 0.3*cm))

    if added:
        add_data = [["#", "Equipment / Équipement"]]
        for i, eq in enumerate(sorted(added), 1):
            add_data.append([str(i), str(eq)])
        add_table = Table(add_data, colWidths=[1.5*cm, 14.5*cm])
        add_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), green),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("BACKGROUND",   (0,1), (-1,-1), green_bg),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[green_bg, colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.5, mid_gray),
            ("TOPPADDING",   (0,0), (-1,-1), 6),
            ("BOTTOMPADDING",(0,0), (-1,-1), 6),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("ALIGN",        (0,0), (0,-1), "CENTER"),
        ]))
        story.append(add_table)
    else:
        story.append(Paragraph("✅ No equipment added. / Aucun équipement ajouté.", style_body))

    # ── DELETED ──────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("➖ Deleted Equipment / Équipements Supprimés", style_section_title))
    story.append(HRFlowable(width="100%", thickness=1.5, color=red))
    story.append(Spacer(1, 0.3*cm))

    if deleted:
        del_data = [["#", "Equipment / Équipement"]]
        for i, eq in enumerate(sorted(deleted), 1):
            del_data.append([str(i), str(eq)])
        del_table = Table(del_data, colWidths=[1.5*cm, 14.5*cm])
        del_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), red),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[red_bg, colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.5, mid_gray),
            ("TOPPADDING",   (0,0), (-1,-1), 6),
            ("BOTTOMPADDING",(0,0), (-1,-1), 6),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("ALIGN",        (0,0), (0,-1), "CENTER"),
        ]))
        story.append(del_table)
    else:
        story.append(Paragraph("✅ No equipment deleted. / Aucun équipement supprimé.", style_body))

    # ═══════════════════════════════════════════════════════
    # PAGE 3 — MODIFIED EQUIPMENT
    # ═══════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("✏️ Modified Parameters / Paramètres Modifiés", style_section_title))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#d68910")))
    story.append(Spacer(1, 0.3*cm))

    if modified:
        mod_data = [["Equipment", "Parameter / Paramètre", "Old Value / Ancienne", "New Value / Nouvelle"]]
        for equip, changes in sorted(modified.items()):
            for param, vals in sorted(changes.items()):
                mod_data.append([
                    str(equip).replace("\n", " ").strip(),
                    str(param).replace("\n", " ").strip(),
                    str(vals["old"]),
                    str(vals["new"]),
                ])

        mod_table = Table(mod_data, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
        mod_table.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#d68910")),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[yellow_bg, colors.white]),
            ("BACKGROUND",    (2,1), (2,-1), colors.HexColor("#fde8e8")),
            ("BACKGROUND",    (3,1), (3,-1), colors.HexColor("#e8f8e8")),
            ("GRID",          (0,0), (-1,-1), 0.5, mid_gray),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("WORDWRAP",      (0,0), (-1,-1), True),
        ]))
        story.append(mod_table)
    else:
        story.append(Paragraph("✅ No parameters modified. / Aucun paramètre modifié.", style_body))

    # ── Footer ───────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=accent_red))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(f"Smart PDF Comparator  |  Generated on {now_str}  |  Section: {section}", style_footer))

    doc.build(story)
    buffer.seek(0)
    return buffer


# =============================
# 📊 EXPORT EXCEL COLORÉ
# =============================
def export_excel_colored(df_new, df_old, key_col, section):
    """
    Compare TOUTES les colonnes présentes dans Data Tables.
    Coloration ligne par ligne :
      - Vert  : équipement ajouté (dans new, pas dans old)
      - Rouge : équipement supprimé (dans old, pas dans new)
      - Jaune : au moins une colonne a changé
      - Blanc : identique
    """
    df_new = df_new.fillna("").astype(str)
    df_old = df_old.fillna("").astype(str)

    # Construire les sets depuis key_col
    new_keys = set(df_new[key_col].str.strip())
    old_keys = set(df_old[key_col].str.strip())

    added   = new_keys - old_keys
    deleted = old_keys - new_keys
    common  = new_keys & old_keys

    # Index old par key
    old_indexed = df_old.set_index(df_old[key_col].str.strip())

    # Colonnes communes pour comparaison
    common_cols = [c for c in df_new.columns if c in df_old.columns and c != key_col]

    # Déterminer les modifiés parmi les communs
    modified = set()
    for _, row in df_new.iterrows():
        k = str(row[key_col]).strip()
        if k in common and k in old_indexed.index:
            old_row = old_indexed.loc[k]
            if isinstance(old_row, pd.DataFrame):
                old_row = old_row.iloc[0]
            for col in common_cols:
                if str(row.get(col, "")).strip() != str(old_row.get(col, "")).strip():
                    modified.add(k)
                    break

    # Styles
    fill_green  = PatternFill("solid", fgColor="C6EFCE")
    fill_red    = PatternFill("solid", fgColor="FFC7CE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_white  = PatternFill("solid", fgColor="FFFFFF")
    fill_header = PatternFill("solid", fgColor="1F3864")

    font_white = Font(bold=True, color="FFFFFF")
    font_bold  = Font(bold=True)
    center     = Alignment(horizontal="center", vertical="center")
    border     = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Result"

    # ── TITRE ────────────────────────────────────────────────
    nb_cols = len(df_new.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws.cell(row=1, column=1, value=f"COMPARAISON - {section}").fill = fill_header
    ws.cell(row=1, column=1).font      = Font(bold=True, color="FFFFFF", size=13)
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).border    = border

    # ── LÉGENDE ──────────────────────────────────────────────
    legends = [
        (fill_green,  "Vert",   "Equipement AJOUTE dans la NEW list"),
        (fill_red,    "Rouge",  "Equipement SUPPRIME (present dans OLD, absent dans NEW)"),
        (fill_yellow, "Jaune",  "Equipement MODIFIE (au moins un parametre a change)"),
        (fill_white,  "Blanc",  "Equipement IDENTIQUE (aucun changement)"),
    ]
    for i, (fill, color_label, meaning) in enumerate(legends, start=2):
        ws.cell(row=i, column=1, value=color_label).fill = fill
        ws.cell(row=i, column=1).font      = font_bold
        ws.cell(row=i, column=1).border    = border
        ws.cell(row=i, column=1).alignment = center
        ws.cell(row=i, column=2, value=meaning).fill = fill
        ws.cell(row=i, column=2).border    = border
        if nb_cols > 2:
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=nb_cols)

    ws.append([])  # ligne vide

    # ── EN-TÊTES ─────────────────────────────────────────────
    header_row = 7
    cols = list(df_new.columns)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.fill      = fill_header
        cell.font      = font_white
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(j)].width = max(18, len(str(col)) + 4)

    # ── DONNÉES NEW (colorées cellule par cellule) ──────────
    for _, row in df_new.iterrows():
        key_val = str(row.get(key_col, "")).strip()

        data_row = [str(row[col]) for col in cols]
        ws.append(data_row)
        cur = ws.max_row

        if key_val in added:
            # Toute la ligne verte
            for j in range(1, len(cols) + 1):
                ws.cell(row=cur, column=j).fill      = fill_green
                ws.cell(row=cur, column=j).border    = border
                ws.cell(row=cur, column=j).alignment = Alignment(vertical="center")
        else:
            # Coloration cellule par cellule
            old_row = None
            if key_val in old_indexed.index:
                old_row = old_indexed.loc[key_val]
                if isinstance(old_row, pd.DataFrame):
                    old_row = old_row.iloc[0]

            for j, col in enumerate(cols, start=1):
                cell = ws.cell(row=cur, column=j)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

                if old_row is not None and col != key_col:
                    new_val = str(row.get(col, "")).strip()
                    old_val = str(old_row.get(col, "")).strip()
                    if new_val != old_val:
                        cell.fill = fill_yellow  # Cellule modifiée
                    else:
                        cell.fill = fill_white   # Identique
                else:
                    cell.fill = fill_white

    # ── ÉQUIPEMENTS SUPPRIMÉS (en rouge à la fin) ─────────────
    if deleted:
        # Ligne séparatrice
        ws.append([""] * len(cols))
        cur = ws.max_row
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(cols))
        ws.cell(row=cur, column=1, value="⬇ EQUIPEMENTS SUPPRIMES (présents dans OLD, absents dans NEW)").fill = fill_red
        ws.cell(row=cur, column=1).font      = Font(bold=True, color="9C0006")
        ws.cell(row=cur, column=1).alignment = center
        ws.cell(row=cur, column=1).border    = border

        for del_key in sorted(deleted):
            # Récupérer la ligne complète depuis old
            if del_key in old_indexed.index:
                old_row = old_indexed.loc[del_key]
                if isinstance(old_row, pd.DataFrame):
                    old_row = old_row.iloc[0]
                data_row = [str(old_row.get(col, "")) for col in cols]
            else:
                data_row = [del_key] + [""] * (len(cols) - 1)

            ws.append(data_row)
            cur = ws.max_row
            for j in range(1, len(cols) + 1):
                ws.cell(row=cur, column=j).fill   = fill_red
                ws.cell(row=cur, column=j).border = border
                ws.cell(row=cur, column=j).font   = Font(italic=True, color="9C0006")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =============================
# 🎨 DISPLAY COMPARISON RESULTS
# =============================
def display_comparison_results(result, section):

    added    = result["added"]
    deleted  = result["deleted"]
    modified = result["modified"]
    total    = len(added) + len(deleted) + len(modified)

    st.markdown("---")
    st.subheader("📊 KPI Summary")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("➕ Added",         len(added),
              delta=f"+{len(added)}" if added else None,
              delta_color="normal")
    c2.metric("➖ Deleted",       len(deleted),
              delta=f"-{len(deleted)}" if deleted else None,
              delta_color="inverse")
    c3.metric("✏️ Modified",      len(modified))
    c4.metric("🔢 Total changes", total)

    st.markdown("---")

    with st.expander(f"➕ Added equipment — {len(added)} item(s)",
                     expanded=len(added) > 0):
        if added:
            df_add = pd.DataFrame({"#": range(1, len(added)+1),
                                   "Equipment": sorted(added)})
            st.dataframe(df_add, use_container_width=True, hide_index=True)
        else:
            st.success("No equipment added.")

    with st.expander(f"➖ Deleted equipment — {len(deleted)} item(s)",
                     expanded=len(deleted) > 0):
        if deleted:
            df_del = pd.DataFrame({"#": range(1, len(deleted)+1),
                                   "Equipment": sorted(deleted)})
            st.dataframe(df_del, use_container_width=True, hide_index=True)
        else:
            st.success("No equipment deleted.")

    with st.expander(f"✏️ Modified parameters — {len(modified)} equipment",
                     expanded=True):
        if not modified:
            st.success("No parameter changes detected.")
            return

        rows = []
        for equip, changes in modified.items():
            for param, vals in changes.items():
                rows.append({
                    "Equipment": equip.replace("\n", " ").strip(),
                    "Parameter": param.replace("\n", " ").strip(),
                    "Old Value": vals["old"],
                    "New Value": vals["new"],
                })

        df_mod = (pd.DataFrame(rows)
                    .sort_values(["Equipment", "Parameter"])
                    .reset_index(drop=True))
        df_mod.index += 1

        st.session_state["df_mod"] = df_mod
        st.session_state["section"] = section

        def color_row(row):
            return [
                "",
                "",
                "background-color:#fff3cd; color:#7d5a00",
                "background-color:#d4edda; color:#155724",
            ]

        st.markdown("#### 🔍 Filter by parameter")
        all_params = sorted(df_mod["Parameter"].unique())
        chosen = st.multiselect(
            "Select parameter(s) to isolate",
            all_params,
            default=[],
            key="param_filter"
        )

        df_display = df_mod.copy()
        if chosen:
            df_display = df_mod[df_mod["Parameter"].isin(chosen)].reset_index(drop=True)
            df_display.index += 1

        st.dataframe(
            df_display.style.apply(color_row, axis=1),
            use_container_width=True,
            hide_index=False,
            height=min(600, 40 + len(df_display) * 36),
        )

        st.caption(f"Showing {len(df_display)} of {len(df_mod)} changes")

        csv = df_mod.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="⬇️ Download all changes as CSV",
            data=csv,
            file_name=f"changes_{section.replace(' ', '_')}.csv",
            mime="text/csv",
        )


# =============================
# 🧭 UI
# =============================
st.sidebar.title("📌 Menu")
menu = st.sidebar.radio("", [
    "🏠 Home",
    "⚡ Quick Compare",
    "📄 Viewer",
    "📑 Index",
    "📊 Tables",
    "🔬 Deep Comparison",
    "📊 Dashboard",
    "📁 Exports & Reports"
])

new_pdf = st.sidebar.file_uploader("🆕 New List", type=["pdf"])
old_pdf = st.sidebar.file_uploader("📁 Old List", type=["pdf"])

# ── Load PDFs once into session_state ─────────────────────
if new_pdf is not None:
    new_name = new_pdf.name
    if st.session_state.get("new_pdf_name") != new_name:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t:
            t.write(new_pdf.read())
            st.session_state["p1"] = t.name
        st.session_state["new_pdf_name"] = new_name
        # Reset all cached data when PDF changes
        for k in ["idx_new", "common_sections", "fnew_current", "fold_current",
                  "last_all_cols", "deep_result", "fnew_snapshot", "fold_snapshot"]:
            st.session_state.pop(k, None)

if old_pdf is not None:
    old_name = old_pdf.name
    if st.session_state.get("old_pdf_name") != old_name:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t:
            t.write(old_pdf.read())
            st.session_state["p2"] = t.name
        st.session_state["old_pdf_name"] = old_name
        for k in ["idx_old", "common_sections", "fnew_current", "fold_current",
                  "last_all_cols", "deep_result", "fnew_snapshot", "fold_snapshot"]:
            st.session_state.pop(k, None)

# Extract index once and cache in session_state
if "p1" in st.session_state and "idx_new" not in st.session_state:
    idx_new, _, _ = extract_index_and_info(st.session_state["p1"])
    st.session_state["idx_new"] = idx_new

if "p2" in st.session_state and "idx_old" not in st.session_state:
    idx_old, _, _ = extract_index_and_info(st.session_state["p2"])
    st.session_state["idx_old"] = idx_old

if "idx_new" in st.session_state and "idx_old" in st.session_state and "common_sections" not in st.session_state:
    idx_new = st.session_state["idx_new"]
    idx_old = st.session_state["idx_old"]
    st.session_state["common_sections"] = sorted(
        list(set(filter_sections(idx_new)) & set(filter_sections(idx_old)))
    )

# =============================
# =============================
# 🏠 HOME
# =============================
if menu == "🏠 Home":

    # ── Hero Header ──────────────────────────────────────────
    st.markdown("""
    <div style="
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 3rem 2.5rem 2rem;
        border-radius: 20px;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,0,0,0.3);
    ">
        <div style="font-size:3.5rem; margin-bottom:0.5rem;">🏭</div>
        <h1 style="color:#e94560; margin:0; font-size:2.8rem; font-weight:800; letter-spacing:1px;">
            Smart PDF Comparator
        </h1>
        <p style="color:#a8b2d8; margin:0.6rem 0 0; font-size:1.1rem;">
            Industrial Electrical Load List — Comparison & Analysis Platform
        </p>
        <p style="color:#6c757d; margin:0.8rem 0 0; font-size:0.85rem;">
            Powered by pdfplumber · pandas · Streamlit
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ── How to use title ─────────────────────────────────────
    st.markdown("""
    <div style="text-align:center; margin-bottom:1.5rem;">
        <h2 style="color:#0f3460; font-size:1.6rem; font-weight:700;">
            📖 How to Use This Application
        </h2>
        <p style="color:#6c757d; font-size:0.95rem;">
            Follow the steps below to compare your PDF load lists efficiently
        </p>
    </div>
    """, unsafe_allow_html=True)

    # ── Step cards ───────────────────────────────────────────
    steps = [
        {
            "icon": "📂",
            "step": "Step 1",
            "title": "Upload PDFs",
            "title_fr": "Charger les PDFs",
            "desc": "Upload your NEW and OLD load list PDF files using the sidebar on the left.",
            "desc_fr": "Chargez vos fichiers PDF (nouvelle et ancienne liste) dans la barre latérale.",
            "color": "#1a5276",
            "bg": "#eaf2fb"
        },
        {
            "icon": "⚡",
            "step": "Step 2",
            "title": "Quick Compare",
            "title_fr": "Comparaison Rapide",
            "desc": "Get an instant overview — are the PDFs identical or different? See similarity percentage.",
            "desc_fr": "Obtenez un aperçu rapide : les PDFs sont-ils identiques ? Voir le pourcentage de similarité.",
            "color": "#7d6608",
            "bg": "#fef9e7"
        },
        {
            "icon": "📑",
            "step": "Step 3",
            "title": "Check Index",
            "title_fr": "Vérifier l'Index",
            "desc": "View the table of contents extracted from each PDF to verify sections.",
            "desc_fr": "Consultez la table des matières extraite de chaque PDF pour vérifier les sections.",
            "color": "#1e8449",
            "bg": "#eafaf1"
        },
        {
            "icon": "📊",
            "step": "Step 4",
            "title": "View Tables",
            "title_fr": "Voir les Tableaux",
            "desc": "Browse the extracted data tables side by side — NEW list vs OLD list.",
            "desc_fr": "Parcourez les tableaux de données extraits côte à côte — liste NEW vs OLD.",
            "color": "#6c3483",
            "bg": "#f5eef8"
        },
        {
            "icon": "🔬",
            "step": "Step 5",
            "title": "Deep Comparison",
            "title_fr": "Comparaison Approfondie",
            "desc": "Select a section, key column and parameters to compare. Find exactly what changed.",
            "desc_fr": "Sélectionnez une section, colonne clé et paramètres à comparer. Trouvez exactement ce qui a changé.",
            "color": "#922b21",
            "bg": "#fdf2f2"
        },
        {
            "icon": "📈",
            "step": "Step 6",
            "title": "Dashboard",
            "title_fr": "Tableau de Bord",
            "desc": "View visual charts and KPIs summarizing all detected changes after comparison.",
            "desc_fr": "Visualisez les graphiques et KPIs résumant tous les changements détectés.",
            "color": "#0f3460",
            "bg": "#eaf2fb"
        },
        {
            "icon": "📁",
            "step": "Step 7",
            "title": "Exports & Reports",
            "title_fr": "Exports et Rapports",
            "desc": "Download a colored Excel file and a bilingual PDF report of all changes.",
            "desc_fr": "Téléchargez un fichier Excel coloré et un rapport PDF bilingue de tous les changements.",
            "color": "#1e8449",
            "bg": "#eafaf1"
        },
    ]

    # Display 2 cards per row then 3 per row
    cols_row1 = st.columns(2)
    cols_row2 = st.columns(3)
    cols_row3 = st.columns(2)

    card_template = """
    <div style="
        background: {bg};
        border-radius: 16px;
        padding: 1.4rem 1.2rem;
        border-left: 5px solid {color};
        box-shadow: 0 2px 12px rgba(0,0,0,0.07);
        height: 100%;
        margin-bottom: 1rem;
    ">
        <div style="font-size:2.2rem; margin-bottom:0.4rem;">{icon}</div>
        <div style="
            background:{color};
            color:white;
            font-size:0.7rem;
            font-weight:700;
            padding:2px 8px;
            border-radius:20px;
            display:inline-block;
            margin-bottom:0.5rem;
            letter-spacing:1px;
        ">{step}</div>
        <h4 style="color:{color}; margin:0.3rem 0 0.1rem; font-size:1rem; font-weight:700;">
            {title}
        </h4>
        <p style="color:#6c757d; font-size:0.72rem; font-style:italic; margin:0 0 0.5rem;">
            {title_fr}
        </p>
        <p style="color:#444; font-size:0.82rem; margin:0 0 0.3rem; line-height:1.5;">
            {desc}
        </p>
        <p style="color:#888; font-size:0.75rem; margin:0; font-style:italic; line-height:1.4;">
            {desc_fr}
        </p>
    </div>
    """

    # Row 1: Step 1 & 2
    for i, col in enumerate(cols_row1):
        s = steps[i]
        col.markdown(card_template.format(**s), unsafe_allow_html=True)

    # Row 2: Step 3, 4, 5
    for i, col in enumerate(cols_row2):
        s = steps[i+2]
        col.markdown(card_template.format(**s), unsafe_allow_html=True)

    # Row 3: Step 6 & 7
    for i, col in enumerate(cols_row3):
        s = steps[i+5]
        col.markdown(card_template.format(**s), unsafe_allow_html=True)

    # ── Quick Status ─────────────────────────────────────────
    st.markdown("---")
    st.markdown("""
    <h3 style="color:#0f3460; font-size:1.2rem; font-weight:700; margin-bottom:1rem;">
        📌 Current Session Status
    </h3>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        if new_pdf:
            st.success(f"🆕 NEW PDF: `{new_pdf.name}`")
        else:
            st.warning("🆕 NEW PDF: Not uploaded yet")
    with c2:
        if old_pdf:
            st.success(f"📁 OLD PDF: `{old_pdf.name}`")
        else:
            st.warning("📁 OLD PDF: Not uploaded yet")
    with c3:
        if st.session_state.get("deep_result"):
            section = st.session_state.get("deep_section", "—")
            n = len(st.session_state["deep_result"]["added"]) + len(st.session_state["deep_result"]["deleted"]) + len(st.session_state["deep_result"]["modified"])
            st.success(f"🔬 Comparison done: **{n}** changes in `{section}`")
        else:
            st.info("🔬 No comparison run yet")

    # ── Footer ───────────────────────────────────────────────
    st.markdown("---")
    st.markdown("""
    <div style="text-align:center; color:#6c757d; font-size:0.8rem; padding:0.5rem 0;">
        🏭 Smart PDF Comparator &nbsp;|&nbsp;
        Powered by pdfplumber · pandas · Streamlit
    </div>
    """, unsafe_allow_html=True)
# =============================
# ⚡ QUICK COMPARE
# =============================
elif menu == "⚡ Quick Compare":
    st.title("⚡ Quick Comparison")

    if new_pdf and old_pdf:

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t1:
            t1.write(new_pdf.read())
            p1 = t1.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t2:
            t2.write(old_pdf.read())
            p2 = t2.name

        idx_new, _, _ = extract_index_and_info(p1)
        idx_old, _, _ = extract_index_and_info(p2)

        same, info = quick_compare(idx_new, idx_old, p1, p2)

        if same:
            st.success("✅ PDFs are IDENTICAL")
            st.metric("Similarity (%)", info["similarity"])
            st.metric("Difference (%)", info["difference"])
        else:
            st.error("❌ PDFs are DIFFERENT")
            st.write("📌 Reason:", info["reason"])
            if "different_pages" in info:
                st.write("📄 Pages:", info["different_pages"])

# =============================
# 📄 VIEWER
# =============================
elif menu == "📄 Viewer":
    col1, col2 = st.columns(2)

    if new_pdf:
        with col1:
            st.subheader("🆕 NEW")
            show_pdf(new_pdf)

    if old_pdf:
        with col2:
            st.subheader("📁 OLD")
            show_pdf(old_pdf)

# =============================
# 📑 INDEX
# =============================
elif menu == "📑 Index":

    if new_pdf:
        with tempfile.NamedTemporaryFile(delete=False) as t:
            t.write(new_pdf.read())
            p = t.name
        idx, _, name = extract_index_and_info(p)
        st.title(name)
        st.write(filter_sections(idx))

    if old_pdf:
        with tempfile.NamedTemporaryFile(delete=False) as t:
            t.write(old_pdf.read())
            p = t.name
        idx, _, _ = extract_index_and_info(p)
        st.write(filter_sections(idx))

# =============================
# 📊 TABLES
# =============================
elif menu == "📊 Tables":

    st.title("📊 Deep Comparison (Tables)")

    if new_pdf and old_pdf:

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t1:
            t1.write(new_pdf.read())
            p1 = t1.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as t2:
            t2.write(old_pdf.read())
            p2 = t2.name

        idx_new, _, _ = extract_index_and_info(p1)
        idx_old, _, _ = extract_index_and_info(p2)

        common = list(set(filter_sections(idx_new)) & set(filter_sections(idx_old)))

        if common:
            section = st.selectbox("🎯 Section", common)

            with pdfplumber.open(p1) as pdf:
                n1 = len(pdf.pages)
            with pdfplumber.open(p2) as pdf:
                n2 = len(pdf.pages)

            r1 = get_section_ranges(filter_sections(idx_new), n1)
            r2 = get_section_ranges(filter_sections(idx_old), n2)

            s1, e1 = r1[section]
            s2, e2 = r2[section]

            tnew = extract_tables_range(p1, s1, e1)
            told = extract_tables_range(p2, s2, e2)

            fnew = merge_tables_clean(tnew, p1, s1)
            fold = merge_tables_clean(told, p2, s2)

            col1, col2 = st.columns(2)

            with col1:
                st.subheader("🆕 NEW")
                st.dataframe(fnew)

            with col2:
                st.subheader("📁 OLD")
                st.dataframe(fold)

# =============================
# 🔬 DEEP COMPARISON UI
# =============================
elif menu == "🔬 Deep Comparison":

    st.title("🔬 Deep Column Comparison")

    # Initialiser session_state
    for key, default in [
        ("deep_result", None),
        ("deep_section", None),
        ("saved_section", None),
        ("saved_key_col", None),
        ("saved_selected_cols", []),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    pdfs_ready = "p1" in st.session_state and "p2" in st.session_state and "common_sections" in st.session_state

    if not new_pdf or not old_pdf:
        st.warning("⚠️ Please upload both PDFs in the sidebar.")
    elif not pdfs_ready:
        st.info("⏳ Loading PDFs... please wait.")
    else:
        p1     = st.session_state["p1"]
        p2     = st.session_state["p2"]
        idx_new = st.session_state["idx_new"]
        idx_old = st.session_state["idx_old"]
        common  = st.session_state["common_sections"]

        if common:

            # ── STEP 1 : Load section data first ─────────────
            saved_section = st.session_state.get("saved_section")
            section_index = common.index(saved_section) if saved_section in common else 0

            # Section selectbox OUTSIDE form so it triggers load
            section = st.selectbox("🎯 Section", common, index=section_index)

            # Load tables when section changes
            if st.session_state.get("saved_section") != section or "fnew_current" not in st.session_state:
                with st.spinner(f"⏳ Loading section '{section}'..."):
                    with pdfplumber.open(p1) as pdf:
                        n1 = len(pdf.pages)
                    with pdfplumber.open(p2) as pdf:
                        n2 = len(pdf.pages)

                    r1 = get_section_ranges(filter_sections(idx_new), n1)
                    r2 = get_section_ranges(filter_sections(idx_old), n2)

                    s1, e1 = r1[section]
                    s2, e2 = r2[section]

                    tnew = extract_tables_range(p1, s1, e1)
                    told = extract_tables_range(p2, s2, e2)

                    fnew = merge_tables_clean(tnew, p1, s1)
                    fold = merge_tables_clean(told, p2, s2)

                    st.session_state["fnew_current"] = fnew.copy()
                    st.session_state["fold_current"] = fold.copy()
                    st.session_state["saved_section"] = section
                    all_cols = sorted(list(set(fnew.columns) & set(fold.columns)))
                    st.session_state["last_all_cols"] = all_cols

            fnew = st.session_state.get("fnew_current", pd.DataFrame())
            fold = st.session_state.get("fold_current", pd.DataFrame())
            all_cols = st.session_state.get("last_all_cols", [])

            if not fnew.empty and not fold.empty and all_cols:

                # ── STEP 2 : Form for key col + columns ──────
                st.markdown("#### ⚙️ Comparison Settings")

                # Key column OUTSIDE form to update available_cols instantly
                saved_key = st.session_state.get("saved_key_col")
                key_index = all_cols.index(saved_key) if saved_key in all_cols else 0
                key_col = st.selectbox("🔑 Key column", all_cols, index=key_index)
                st.session_state["saved_key_col"] = key_col

                # Columns to compare OUTSIDE form to avoid value lag
                available_cols = [c for c in all_cols if c != key_col]
                saved_cols = st.session_state.get("saved_selected_cols", [])
                default_cols = [c for c in saved_cols if c in available_cols]
                selected_cols = st.multiselect(
                    "📌 Columns to compare (screen results)",
                    available_cols,
                    default=default_cols,
                    key="multiselect_cols"
                )
                st.session_state["saved_selected_cols"] = selected_cols

                st.markdown("*When ready, click **Run Deep Comparison***")
                if st.button("🚀 Run Deep Comparison", use_container_width=True):
                    # Read directly from session_state to avoid lag
                    cols_to_compare = st.session_state.get("multiselect_cols", selected_cols)
                    key_to_use = st.session_state.get("saved_key_col", key_col)
                    with st.spinner("⏳ Comparing data..."):
                        result = deep_compare(fnew, fold, key_to_use, cols_to_compare)
                        st.session_state["deep_result"]  = result
                        st.session_state["deep_section"] = section
                        st.session_state["fnew_snapshot"] = fnew.copy()
                        st.session_state["fold_snapshot"] = fold.copy()
                        st.session_state["excel_cols_snapshot"] = []
                        st.session_state["excel_ready"] = False
                        if "param_filter" in st.session_state:
                            del st.session_state["param_filter"]
                    st.success("✅ Comparison done!")

            if st.session_state.get("deep_result") is not None:
                display_comparison_results(
                    st.session_state["deep_result"],
                    st.session_state["deep_section"]
                )
                st.markdown("---")
                st.info("📁 Go to **Exports & Reports** in the menu to download the colored Excel file.")

# =============================
# 📊 DASHBOARD
# =============================
elif menu == "📊 Dashboard":

    st.markdown("""
    <div style="
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 1.8rem 2rem;
        border-radius: 16px;
        margin-bottom: 1.5rem;
    ">
        <h1 style="color:#e94560; margin:0; font-size:2rem;">📊 Comparison Dashboard</h1>
        <p style="color:#a8b2d8; margin:0.3rem 0 0; font-size:0.95rem;">
            Visual summary of the last deep comparison
        </p>
    </div>
    """, unsafe_allow_html=True)

    deep_result  = st.session_state.get("deep_result", None)
    deep_section = st.session_state.get("deep_section", "—")

    if deep_result is None:
        st.markdown("""
        <div style="
            background:#f8f9fa; border-radius:16px; padding:3rem;
            text-align:center; border:2px dashed #dee2e6;
        ">
            <div style="font-size:3rem;">🔬</div>
            <h3 style="color:#6c757d;">No comparison available yet</h3>
            <p style="color:#adb5bd;">
                Go to <b>Deep Comparison</b>, run an analysis,
                then come back here to see the dashboard.
            </p>
        </div>
        """, unsafe_allow_html=True)
    else:
        n_added    = len(deep_result["added"])
        n_deleted  = len(deep_result["deleted"])
        n_modified = len(deep_result["modified"])
        n_total    = n_added + n_deleted + n_modified

        # ── Section badge ────────────────────────────────────
        st.markdown(f"""
        <div style="
            background:#eaf2fb; border-radius:10px; padding:0.6rem 1rem;
            border-left:4px solid #1a5276; margin-bottom:1rem;
            display:inline-block;
        ">
            📋 Section analysed: <b>{deep_section}</b>
        </div>
        """, unsafe_allow_html=True)

        # ── KPI Cards ────────────────────────────────────────
        k1, k2, k3, k4 = st.columns(4)

        def kpi_card(col, emoji, label, label_fr, value, bg, fg):
            col.markdown(f"""
            <div style="
                background:{bg}; border-radius:14px; padding:1.4rem 1rem;
                text-align:center; border:1px solid {fg}33;
                box-shadow:0 2px 8px rgba(0,0,0,0.06);
            ">
                <div style="font-size:2rem;">{emoji}</div>
                <div style="font-size:2.2rem; font-weight:800; color:{fg}; margin:0.3rem 0;">{value}</div>
                <div style="font-size:0.85rem; font-weight:600; color:{fg};">{label}</div>
                <div style="font-size:0.72rem; color:#888; font-style:italic;">{label_fr}</div>
            </div>
            """, unsafe_allow_html=True)

        kpi_card(k1, "➕", "Added",    "Ajoutés",    n_added,    "#eafaf1", "#1e8449")
        kpi_card(k2, "➖", "Deleted",  "Supprimés",  n_deleted,  "#fdf2f2", "#c0392b")
        kpi_card(k3, "✏️", "Modified", "Modifiés",   n_modified, "#fef9e7", "#d68910")
        kpi_card(k4, "🔢", "Total",    "Total",      n_total,    "#eaf2fb", "#1a5276")

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Charts ───────────────────────────────────────────
        modified = deep_result["modified"]
        rows_mod = []
        for equip, changes in modified.items():
            for param, vals in changes.items():
                rows_mod.append({
                    "Equipment": equip.replace("\n", " ").strip(),
                    "Parameter": param.replace("\n", " ").strip(),
                    "Old": vals["old"],
                    "New": vals["new"],
                })
        df_dash = pd.DataFrame(rows_mod) if rows_mod else pd.DataFrame()

        if not df_dash.empty:
            import json
            chart_col1, chart_col2 = st.columns(2)

            with chart_col1:
                st.markdown("""
                <div style="background:#f8f9fa; border-radius:12px; padding:1rem; text-align:center;">
                <h4 style="color:#0f3460; margin-bottom:0.5rem;">🍩 Change Breakdown</h4>
                """, unsafe_allow_html=True)
                donut_labels = ["Added", "Deleted", "Modified"]
                donut_values = [n_added, n_deleted, n_modified]
                donut_colors = ["#1e8449", "#c0392b", "#d68910"]
                donut_html = f"""
                <canvas id="donut" width="280" height="260"></canvas>
                <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
                <script>
                new Chart(document.getElementById('donut'), {{
                    type: 'doughnut',
                    data: {{
                        labels: {json.dumps(donut_labels)},
                        datasets: [{{
                            data: {json.dumps(donut_values)},
                            backgroundColor: {json.dumps(donut_colors)},
                            borderWidth: 3,
                            borderColor: '#fff'
                        }}]
                    }},
                    options: {{
                        cutout: '65%',
                        plugins: {{
                            legend: {{ position: 'bottom', labels: {{ font: {{ size: 12 }} }} }},
                        }}
                    }}
                }});
                </script>
                """
                st.components.v1.html(donut_html, height=300)
                st.markdown("</div>", unsafe_allow_html=True)

            with chart_col2:
                st.markdown("""
                <div style="background:#f8f9fa; border-radius:12px; padding:1rem; text-align:center;">
                <h4 style="color:#0f3460; margin-bottom:0.5rem;">📊 Top Modified Parameters</h4>
                """, unsafe_allow_html=True)
                param_counts = df_dash["Parameter"].value_counts().head(8)
                bar_labels   = param_counts.index.tolist()
                bar_values   = param_counts.values.tolist()
                bar_html = f"""
                <canvas id="bar" width="380" height="260"></canvas>
                <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
                <script>
                new Chart(document.getElementById('bar'), {{
                    type: 'bar',
                    data: {{
                        labels: {json.dumps(bar_labels)},
                        datasets: [{{
                            label: 'Changes',
                            data: {json.dumps(bar_values)},
                            backgroundColor: '#0f3460cc',
                            borderColor: '#e94560',
                            borderWidth: 1,
                            borderRadius: 6
                        }}]
                    }},
                    options: {{
                        indexAxis: 'y',
                        plugins: {{ legend: {{ display: false }} }},
                        scales: {{
                            x: {{ beginAtZero: true, ticks: {{ stepSize: 1 }} }},
                            y: {{ ticks: {{ font: {{ size: 10 }} }} }}
                        }}
                    }}
                }});
                </script>
                """
                st.components.v1.html(bar_html, height=300)
                st.markdown("</div>", unsafe_allow_html=True)

            # ── Top 5 most impacted equipment ────────────────
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("""
            <h4 style="color:#0f3460; margin-bottom:0.8rem;">
                🏆 Top 5 Most Impacted Equipment
            </h4>
            """, unsafe_allow_html=True)

            top5 = (df_dash.groupby("Equipment")
                           .size()
                           .sort_values(ascending=False)
                           .head(5)
                           .reset_index())
            top5.columns = ["Equipment", "# Parameters Changed"]
            top5.index += 1

            medal = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
            top5.insert(0, "Rank", medal[:len(top5)])

            colors_top = ["#fff9c4","#fff3cd","#fdebd0","#fdf2f8","#f0f4ff"]
            def highlight_top(row):
                idx = row.name - 1
                c = colors_top[idx] if idx < len(colors_top) else ""
                return [f"background-color:{c}"]*len(row)

            st.dataframe(
                top5.style.apply(highlight_top, axis=1),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("No modified parameters to display in charts.")

        # ── Footer ───────────────────────────────────────────
        st.markdown("---")
        st.markdown("""
        <div style="text-align:center; color:#6c757d; font-size:0.8rem;">
            💡 Go to <b>Exports & Reports</b> to download Excel and PDF report
        </div>
        """, unsafe_allow_html=True)


# =============================
# 📁 EXPORTS & REPORTS
# =============================
elif menu == "📁 Exports & Reports":

    st.title("📁 Exports & Reports")

    deep_result  = st.session_state.get("deep_result", None)
    deep_section = st.session_state.get("deep_section", None)
    fnew_snap    = st.session_state.get("fnew_snapshot", None)
    fold_snap    = st.session_state.get("fold_snapshot", None)
    key_col_snap = st.session_state.get("saved_key_col", None)

    if deep_result is None or fnew_snap is None:
        st.warning("⚠️ No comparison has been run yet. Please go to **🔬 Deep Comparison** first and run a comparison.")
    else:
        st.success(f"✅ Comparison available — Section: **{deep_section}**")

        # ── KPI summary ──────────────────────────────────────
        n_added    = len(deep_result["added"])
        n_deleted  = len(deep_result["deleted"])
        n_modified = len(deep_result["modified"])
        n_total    = n_added + n_deleted + n_modified

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("➕ Added",    n_added)
        c2.metric("➖ Deleted",  n_deleted)
        c3.metric("✏️ Modified", n_modified)
        c4.metric("🔢 Total",    n_total)

        st.markdown("---")

        # ── Excel Export ──────────────────────────────────────
        st.subheader("📊 Colored Excel Export")

        st.markdown("""
        <div style="background:#f0f7ff; border-radius:10px; padding:0.8rem 1rem; margin-bottom:1rem; border-left: 4px solid #1a5276;">
            <b>Color legend:</b><br>
            🟢 <b>Green</b> = Equipment ADDED &nbsp;|&nbsp;
            🔴 <b>Red</b> = Equipment DELETED &nbsp;|&nbsp;
            🟡 <b>Yellow cell</b> = Modified value &nbsp;|&nbsp;
            ⬜ <b>White</b> = Identical
        </div>
        """, unsafe_allow_html=True)

        # ── FORM : select columns then generate ──────────────
        with st.form("excel_export_form"):
            st.markdown("*Select the columns to include in the Excel file, then click **Generate Excel***")
            all_available = [c for c in fnew_snap.columns if c != key_col_snap]
            saved_excel_cols = st.session_state.get("saved_excel_cols", [])
            default_excel = [c for c in saved_excel_cols if c in all_available]

            excel_cols = st.multiselect(
                "📋 Select columns to include in Excel",
                all_available,
                default=default_excel if default_excel else all_available,
                help="Select exactly the columns you want in the exported Excel file"
            )
            generate = st.form_submit_button("⚙️ Generate Excel", use_container_width=True)

        if generate:
            st.session_state["saved_excel_cols"] = excel_cols
            st.session_state["excel_ready"] = True

        excel_cols = st.session_state.get("saved_excel_cols", all_available if "all_available" in dir() else [])

        if st.session_state.get("excel_ready") and excel_cols:
            with st.spinner("⏳ Generating Excel file..."):
                excel_cols_final = [key_col_snap] + [c for c in excel_cols if c != key_col_snap]
                fnew_excel = fnew_snap[[c for c in excel_cols_final if c in fnew_snap.columns]]
                fold_excel = fold_snap[[c for c in excel_cols_final if c in fold_snap.columns]]

                excel_buffer = export_excel_colored(
                    fnew_excel,
                    fold_excel,
                    key_col_snap,
                    deep_section
                )

                import datetime
                now = datetime.datetime.now().strftime("%Y%m%d_%H%M")
                filename = f"comparison_{deep_section.replace(' ', '_')}_{now}.xlsx"

            st.success("✅ Excel file ready!")
            st.download_button(
                label="⬇️ Download Colored Excel",
                data=excel_buffer,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.caption(f"File: `{filename}` — {len(excel_cols_final)} columns — {len(fnew_excel)} rows")
        elif not st.session_state.get("excel_ready"):
            st.info("👆 Select your columns above and click **Generate Excel**")
        else:
            st.warning("Please select at least one column to export.")

        st.markdown("---")

        # ── PDF Report ───────────────────────────────────────
        st.subheader("📝 PDF Report / Rapport PDF")

        st.markdown("""
        <div style="background:#fef9e7; border-radius:10px; padding:0.8rem 1rem; margin-bottom:1rem; border-left: 4px solid #d68910;">
            The report contains / Le rapport contient :<br>
            📋 <b>Cover page</b> with KPI summary &nbsp;|&nbsp;
            ➕ <b>Added equipment</b> &nbsp;|&nbsp;
            ➖ <b>Deleted equipment</b> &nbsp;|&nbsp;
            ✏️ <b>Modified parameters</b> (Old → New)
        </div>
        """, unsafe_allow_html=True)

        if st.button("⚙️ Generate PDF Report", use_container_width=True):
            with st.spinner("⏳ Generating PDF report..."):
                key_col_for_pdf = st.session_state.get("saved_key_col", "—")
                pdf_buffer = generate_pdf_report(
                    deep_result,
                    deep_section,
                    key_col_for_pdf
                )
                st.session_state["pdf_buffer"]  = pdf_buffer
                st.session_state["pdf_ready"]   = True
                st.session_state["pdf_filename"] = f"report_{deep_section.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            st.success("✅ PDF Report ready!")

        if st.session_state.get("pdf_ready"):
            st.download_button(
                label="⬇️ Download PDF Report",
                data=st.session_state["pdf_buffer"],
                file_name=st.session_state["pdf_filename"],
                mime="application/pdf",
                use_container_width=True
            )
            st.caption(f"File: `{st.session_state['pdf_filename']}`")